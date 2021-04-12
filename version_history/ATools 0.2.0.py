""" 
ATools = Analytical Tools
import ATools as at
Use info()
"""
ver = "0.2.0"
author = "sweps91"

def info():
    """Show list of the functions"""

    list_of_funcs = ["tips() = Shortcut tips",
                    "percent_change(list_of_values, rounded=1) = Create the list of the percent change",
                    "percent_change_fl(list_of_values, rounded=1) = Create the list of the percent change between the first and the last item",
                    "percent_change_choice(list_of_values, present_index, past_index, rounded=1) = Create the list of the percent change between two choiced numbers (indexes)",
                    "percent_change_space(list_of_values, space=2, rounded=1) = Create the list of the percent change with specified space [space = 4; past = index 0, present = index 4]",
                    "percent_change_skip(list_of_values, skip, rounded=1) = Create the list of the percent change with skip range [skip = 4; past = index -5, present = index -1]",
                    "percent_change_two_lists = Percent changes between values in two lists",
                    "percent_operation(A, B, A_percent=100, B_percent=False) = Percent operation - find the unknown",
                    "list_rounded(list_of_values) = Creating the new list of the rounded numbers",
                    "modulo(a_num, b_num) = Counting modulo and floor division; a_num devided by b_num",
                    "median(list_of_values) = Median, or 50th percentile, of grouped data",
                    "csv_writer (filename, list_of_labels, *lists_of_values) = Creating csv file",
                    "csv_reader(filename, list_of_columns) =  Reading csv file and making the lists",
                    "list_operator (num_list, operation_num, operator, rounded=1) = num_list[i] operator (default = *) operation_num",
                    "list_to_lists(list_of_values, *name_of_columns) = Transform one data list to more lists",
                    "list_int(list_of_values) = Creating list of integers",
                    "list_float(list_of_values) = Creating list of float nums",
                    "list_sum_division(list_of_values, rounded=1) = Sum of nums divided by len of list",
                    "lists_sum_division(*args, rounded=1) = Take the lists and count average",
                    "list_from_index(list_of_lists, index) = Creating new list from given index in lists",
                    "lists_from_indexes(list_of_lists, index_range) = Create lists in list from indexes (columns)",
                    "list_from_indexes_and_sum_division(list_of_lists, index_range, rounded=1) = Lists_from_indexes + lists_sum_division",
                    "quartal_list(start, end, space=1) = Creating quartal list",
                    "months_cz_list(start, end, space=1) = Creating months CZ list",
                    "months_num_list(start, end, space=1) = Creating months num list",
                    "for_string_in_range(first_str, second_str, start, end, space=1) = Creating desired string with nums",
                    "test_shape(*args, len_num=None) = Testing shape",
                    "list_to_dict_sort_and_back (list_one, list_two, reverse=False) = Take two lists - create dictionary - sort values - create two list sorted according values",
                    "two_lists_difference (List_A, List_B, rounded=1) = Values from List_A minus values from List_B = counting difference",
                    "per_population (values_list, population_list, per_population=1_000, rounded=0) = Values devided per population",
                    "decrease_or_increase_numbers (values_list, operation_number, increase=False, rounded=0) = Numbers in list increased or decreased by operation_number",
                    "analyse_num_list(num_list) = Analysing num list",
                    "clear_string_from_list(list_of_values) = Clear strings from list",
                    "clear_string_from_lists(list_of_lists) = Clear strings from lists",
                    "excel_writer(filename, list_of_labels, list_of_lists, write_lables=True) = Write data to xlsx",
                    "excel_reader(filename, sheet_name, list_feed_column=False, if_cell_none, clear_string=False, len_columns=None) = Read data from xlsx",
                    ]

    list_of_funcs.sort()
    print("\n:: THE FUNCTIONS LIST ::\n")
    for function in list_of_funcs:
        print(function)
    print(f"\nNumber of functions: {len(list_of_funcs)}\n\nATools version: {ver}\nCreated by {author}\n")

    if __name__ == "__main__":
        __ = input("")

def percent_change(list_of_values, rounded=1):
    """Create the list of the percent change
        Growth Rate = ((present - past) / past) * 100 """

    change, working_list = [], []
    working_list.append(list_of_values[0])

    for num in list_of_values[1:]:
        working_list.append(num)
        growth_rate = (((working_list[-1] - working_list[-2]) / working_list[-2]) * 100 )
        growth_rate = (round(growth_rate, rounded))
        if rounded == 0:
            growth_rate = (int(growth_rate))
        change.append(growth_rate)

    return change

def percent_change_fl(list_of_values, rounded=1):
    """Create the list of the percent change between the first and the last item
        Growth Rate = ((present - past) / past) * 100 """

    change = []

    growth_rate = (((list_of_values[-1] - list_of_values[0]) / list_of_values[0]) * 100 )
    growth_rate = (round(growth_rate, rounded))
    if rounded == 0:
        growth_rate = (int(growth_rate))
    change.append(growth_rate)

    return change

def percent_change_choice(list_of_values, present_index, past_index, rounded=1):
    """Create the list of the percent change between two choiced numbers / indexes
        Put indexes of the numbers to the parameters
        Growth Rate = ((present - past) / past) * 100 """

    change = []

    growth_rate = (((list_of_values[present_index] - list_of_values[past_index]) / list_of_values[past_index]) * 100 )
    growth_rate = (round(growth_rate, rounded))
    if rounded == 0:
        growth_rate = (int(growth_rate))
    change.append(growth_rate)

    return change

def percent_change_space(list_of_values, space=2, rounded=1):
    """Create the list of the percent change with specified space [space = 4; past = index 0, present = index 4]
        Growth Rate = ((present - past) / past) * 100 """

    change = []
    present_index = (len(list_of_values)) - 1
    past_index = present_index - space
    
    while past_index >= 0:
        growth_rate = (((list_of_values[present_index] - list_of_values[past_index]) / list_of_values[past_index]) * 100 )
        growth_rate = (round(growth_rate, rounded))
        if rounded == 0:
            growth_rate = (int(growth_rate))
        change.insert(0, growth_rate)
        present_index -= 1 
        past_index -= 1

    return change

def percent_change_skip(list_of_values, skip, rounded=1):
    """Create the list of the percent change with skip range [skip = 4; past = index -5, present = index -1]
        Growth Rate = ((present - past) / past) * 100 """

    change = []
    value = skip
    cursor = skip + 1
    present = -1
    past = -cursor

    while cursor < (len(list_of_values)):
        growth_rate = (((list_of_values[present] - list_of_values[past]) / list_of_values[past]) * 100 )
        growth_rate = (round(growth_rate, rounded))
        if rounded == 0:
            growth_rate = (int(growth_rate))
        change.insert(0, growth_rate)
        cursor += value
        present -= value
        past -= value 

    return change

def percent_change_two_lists (past_list, present_list, rounded=1):
    """Percent changes between values in two lists
    Growth Rate = ((present - past) / past) * 100 """

    data = []
    len_data = (len(past_list))

    for i in range (len_data):
        x = ((present_list[i] - past_list[i]) / past_list[i]) * 100
        x = round(x, rounded)
        if rounded == 0:
            x = (int(x))
        data.append(x)

    return data

def percent_operation(A, B, A_percent=100, B_percent=False):
    """Percent operation - find the unknown"""

    if B_percent==False:
        x = A / A_percent
        y = B / x
        y = (round(y, 2))
        difference = A_percent - y
        difference = (round(difference, 2))
        print(f"{A} is {A_percent} %\n{B} is {y} %\nDifference is {difference} %")

    else:
        x = A / A_percent
        y = x * B
        y = (round(y, 2))
        difference = A_percent - B
        difference = (round(difference, 2))
        print(f"{A} is {A_percent} %\n{y} is {B} %\nDifference is {difference} %")

def list_rounded(list_of_values):
    "Creating the new list of the rounded numbers"

    rounded = []

    for num in list_of_values:
        num = (round(num))
        rounded.append(num)

    return rounded 

def modulo(a_num, b_num):
    """
    Counting modulo and floor division; a_num devided by b_num
    """  

    division = a_num / b_num
    floor_division = a_num // b_num
    floor_division = int(floor_division)
    modulo = a_num % b_num
    
    print(a_num, " devided by ", b_num, " is: ", division)
    print("Floor division / Without remain: ", floor_division)
    print("Remain is / Modulo: ", modulo)

def csv_writer (filename, list_of_labels, list_of_lists):
    """Creating csv file
    using csv module"""

    import csv
  
    with open(filename, mode='w', newline="") as csv_file:
        file_writer = csv.writer(csv_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        #OR: file_writer = csv.writer(csv_file)

        file_writer.writerow(list_of_labels)

        len_data = (len(list_of_lists[0]))
        index = len_data 
        row_range = (len(list_of_lists)) 

        for i in range (index):
            row = []
            for r in range (row_range):
                x = list_of_lists[r][i]
                row.append(x)

            file_writer.writerow(row)

def csv_reader(filename, list_of_columns):
    """Reading csv file and making the lists
    using csv module"""

    import csv

    data = []

    for n in list_of_columns:
        n = []
        data.append(n)

    index = (len(list_of_columns))
    
    with open(filename, newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if (len(row)) == 0:
                pass
            
            else:
                for i in range (index):
                    x = row[i]
                    data[i].append(x)
                                  
    return data

def list_to_lists(list_of_values, num_of_lists):
    """Transform one data list to more lists"""
    
    lol = []
    for n in range(num_of_lists):
        n = []
        lol.append(n)

    space = num_of_lists
    index_cursor = space - 1

    while index_cursor >= 0:
        
        cursor = 1
        loop_cursor = cursor
        current_index = index_cursor
        current_list = index_cursor

        while loop_cursor <= (len(list_of_values)) / num_of_lists:
            x = list_of_values[current_index]
            lol[current_list].append(x)
            current_index += space
            loop_cursor += 1

        index_cursor -= 1

    #reverse lol list
    return lol

def median(list_of_values):
    """Median, or 50th percentile, of grouped data
    using statistics module"""

    import statistics

    median = statistics.median_grouped(list_of_values)

    return median

def list_operator (num_list, operation_num, operator="*", rounded=1):
    """num_list[i] operator (default = *) operation_num"""

    data = []
    len_data = (len(num_list))

    for i in range (len_data):
        if operator == "+":
            x = num_list[i] + operation_num
        elif operator == "-":
            x = num_list[i] - operation_num
        elif operator == "/":
            x = num_list[i] / operation_num
        elif operator == "//":
            x = num_list[i] // operation_num
        elif operator == "%":
            x = num_list[i] % operation_num
        else:
            x = num_list[i] * operation_num
        x = (round(x, rounded))
        if rounded == 0:
            x = (int(x))
        data.append(x)

    return data

def list_int(list_of_values):
    """Creating list of integers
    Float nums will be rounded"""

    values = []
    for l in list_of_values:
        l = (float(l))
        l = (round(l))
        l = (int(l))
        
        values.append(l)

    return values

def list_float(list_of_values):
    """Creating list of float nums"""

    values = []
    for l in list_of_values:
        l = (float(l))
        
        values.append(l)

    return values

def list_sum_division(list_of_values, rounded=1):
    """Sum of nums divided by len of list"""

    sum_nums = (sum(list_of_values))
    len_list = (len(list_of_values))
    x = sum_nums / len_list
    x = (round(x,rounded))
    if rounded == 0:
        x = (int(x))

    return x

def lists_sum_division(*args, rounded=1):
    """Take the lists and count average"""

    li = []
    for l in args:
        x = (len(l))
        y = (sum(l))
        z = y / x
        z = (round(z, rounded))
        if rounded == 0:
            z = (int(z))
        li.append(z)

    return li

def list_from_index(list_of_lists, index):
    """Creating new list from given index in lists"""

    index_list = []
    for l in list_of_lists:
        l = l[index]
        index_list.append(l)

    return index_list

def lists_from_indexes(list_of_lists, index_range):
    """Create lists in list from indexes (columns)"""

    index_list = []

    for n in range (index_range + 1):
        li = []
        for l in list_of_lists:
            l = l[n]
            li.append(l)
        index_list.append(li)

    return index_list

def list_from_indexes_and_sum_division(list_of_lists, index_range, rounded=1):
    """Lists_from_indexes + lists_sum_division"""

    a = lists_from_indexes(list_of_lists, index_range)
    b = []
    c = (len(a))
    for n in range (c):
        x = (len(a[n]))
        y = (sum(a[n]))
        z = y / x
        z = (round(z, rounded))
        if rounded == 0:
            z = (int(z))
        b.append(z)

    return b

def quartal_list(start, end, space=1):
    """Creating quartal list"""

    q_list = []

    for n in range (start, end+1, space):
        n = (str(n))
        q1 = "Q1 " + n
        q2 = "Q2 " + n
        q3 = "Q3 " + n
        q4 = "Q4 " + n
        q_list.append(q1)
        q_list.append(q2)
        q_list.append(q3)
        q_list.append(q4)

    return q_list

def months_cz_list(start, end, space=1):
    "Creating months CZ list"

    cz_list = []

    for n in range(start, end +1, space):
        n = (str(n))
        m1 = "Leden " + n; cz_list.append(m1)
        m2 = "Únor " + n; cz_list.append(m2)
        m3 = "Březen " + n; cz_list.append(m3)
        m4 = "Duben " + n; cz_list.append(m4)
        m5 = "Květen " + n; cz_list.append(m5)
        m6 = "Červen " + n; cz_list.append(m6)
        m7 = "Červenec " + n; cz_list.append(m7)
        m8 = "Srpen " + n; cz_list.append(m8)
        m9 = "Září " + n; cz_list.append(m9)
        m10 = "Říjen " + n; cz_list.append(m10)
        m11 = "Listopad " + n; cz_list.append(m11)
        m12 = "Prosinec " + n; cz_list.append(m12)

    return cz_list

def months_num_list(start, end, space=1):
    "Creating months num list"

    num_list = []

    for n in range(start, end +1, space):
        n = (str(n))
        m1 = "01/" + n; num_list.append(m1)
        m2 = "02/" + n; num_list.append(m2)
        m3 = "03/" + n; num_list.append(m3)
        m4 = "04/" + n; num_list.append(m4)
        m5 = "05/" + n; num_list.append(m5)
        m6 = "06/" + n; num_list.append(m6)
        m7 = "07/" + n;num_list.append(m7)
        m8 = "08/" + n; num_list.append(m8)
        m9 = "09/" + n; num_list.append(m9)
        m10 = "10/" + n; num_list.append(m10)
        m11 = "11/" + n; num_list.append(m11)
        m12 = "12/" + n; num_list.append(m12)

    return num_list

def for_string_in_range(first_str, second_str, start, end, space=1):
    "Creating desired string with nums"

    str_list = []

    for n in range (start, end + 1, space):
        n = (str(n))
        x = first_str + n + second_str
        str_list.append(x)

    return str_list

def test_shape(*args, len_num=None):
    """Testing shape"""

    li = []

    if len_num:
        for l in args:
            len_list = (len(l))
            if len_list == len_num:
                li.append(len_list)

            else: 
                li.append(len_list)
                err = (len(l))
                print(f"{l} = {err}")

    else:
        for l in args:
            l = (len(l))
            li.append(l)

    print(li)

def list_to_dict_sort_and_back (list_one, list_two, reverse=False):
    """Take two lists - create dictionary - sort values - create two list sorted according values"""

    result = [[], []]

    lists_to_dict = dict(zip(list_one, list_two))
    ltd = lists_to_dict

    if reverse != False:
        sorted_dict = sorted(ltd.items(), key = lambda x: x[1], reverse = False)
    else:
        sorted_dict = sorted(ltd.items(), key = lambda x: x[1], reverse = True)

    for sd in sorted_dict:
        key = sd[0]
        value = sd[1]
        result[0].append(key)
        result[1].append(value)

    return result

def two_lists_difference (List_A, List_B, rounded=1):
    """Values from List_A minus values from List_B = counting difference"""

    data = []
    len_data = (len(List_A))

    for i in range (len_data):
        x = List_A[i] - List_B[i]
        x = round(x, rounded)
        if rounded == 0:
            x = (int(x))
        data.append(x)

    return data

def per_population (values_list, population_list, per_population=1_000, rounded=0):
    """Values devided per population"""

    data = []
    len_data = (len(values_list))

    for i in range (len_data):
        x = population_list[i] / per_population
        y = values_list[i] / x  
        y = round(y, rounded)
        if rounded == 0:
            y = (int(y))

        data.append(y)

    return data

def decrease_or_increase_numbers (values_list, operation_number, increase=False, rounded=0):
    """Numbers in list increased or decreased by operation_number
    Division or multiplication"""

    data = []
    len_data = (len(values_list))

    for i in range (len_data):

        if increase != False:
            y = values_list[i] * operation_number  
        else:
            y = values_list[i] / operation_number 

        y = round(y, rounded)

        if rounded == 0:
            y = (int(y))

        data.append(y)

    return data

def analyse_num_list(num_list):
    """Analysing num list"""

    print("-----------------\nANALYSE NUM LIST:\n-----------------")
    a = (len(num_list)); a = (round(a, 2))
    print(f"Len = {a}")
    b = (sum(num_list)); b = round(b, 2)
    print(f"Sum = {b}")
    c = list_sum_division(num_list, rounded=2)
    print(f"Avg = {c}")
    d = median(num_list)
    print(f"Median = {d}")
    print("-----------------")
    sorted_list = (sorted(num_list))
    percent_operation(sorted_list[-1], sorted_list[0])

def clear_string_from_list(list_of_values):
    """Clear strings from list"""

    for i in list_of_values:
            
        try:
            i = (str(i))
            list_of_values.remove(i)
        except ValueError:
            pass
    
    return list_of_values

def clear_string_from_lists(list_of_lists):
    """Clear strings from lists"""

    for li in list_of_lists:
        clear_string_from_list(li)

    return list_of_lists

def excel_writer(filename, list_of_labels, list_of_lists, write_lables=True):
    """Write data to xlsx
    using openpyxl library"""
    from openpyxl import Workbook

    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Data can be assigned directly to cells
    #ws['A1'] = 42

    # Rows can also be appended
    if write_lables==True:
        ws.append(list_of_labels)

    len_data = (len(list_of_lists[0]))
    index = len_data 
    row_range = (len(list_of_lists)) 

    for i in range (index):
        row = []
        for r in range (row_range):
            x = list_of_lists[r][i]
            row.append(x)

        ws.append(row)

    # Save the file
    wb.save(filename)

def excel_reader(filename, sheet_name="Sheet", list_feed_column=False, if_cell_none="empty", clear_string=False, len_columns=None):
    """Read data from xlsx
    if_cell_none = "pass" => pass; if_cell_none = "None" => append None
    using openpyxl library"""

    from openpyxl import load_workbook

    wb = load_workbook(filename)
    ws = wb[sheet_name]

    data = []
          
    row_index = 0
    for row in ws.rows:
        r = []
        data.append(r)
        for cell in row:
            #print(cell.value)
            cv = cell.value
            if cv == None:
                if if_cell_none != "pass" and "None":
                    cv = if_cell_none
                    data[row_index].append(cv)

                elif if_cell_none == "None":
                    cv = None
                    data[row_index].append(cv)

                else:
                    pass

            else:
                data[row_index].append(cv)
        row_index += 1
    
    data_two = []
    if list_feed_column != False:
        len_data = (len(data))

        len_of_columns = (len(data[0]))
        if len_columns != None:
            len_of_columns = len_columns

        for l in range (len_of_columns):
            l = []
            data_two.append(l)

        for li in range (len_data):
                for i in range (len_of_columns):
                    try:
                        data_two[i].append(data[li][i])
                    except IndexError:
                        pass

    if clear_string == True and list_feed_column != False:
        clear_string_from_lists(data_two)
    if clear_string == True and list_feed_column == False:
        clear_string_from_lists(data)
    #   for li in data_two:
    #       for i in li:
    #           try:
    #               i = (str(i))
    #               data_two[data_two.index(li)].remove(i)
    #           except ValueError:
    #               pass

    wb.close()
    if list_feed_column != False:
        return data_two
    else:
        return data













def tips():
    """Shortcut tips"""

    print("lists_to_dict = dict(zip(a, [b, c]))")
    print("dict_to_list = dict.values()")

if __name__ == "__main__":
    info()