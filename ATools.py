""" 
ATools = Analytical Tools
import ATools as at
Use at.info()
"""
ver = "0.8.0"
verdate = "December 2021"
created = "March 2021"
author = "sweps91"
Lines = 1_700

def info():
    """Show list of the functions"""

    list_of_funcs = ["tips() = Shortcut tips",
                    "change(list_of_values, rounded=0) = Count change: index 1 - index 0; i 2 - i 1 etc.",
                    "change_fl(list_of_values, rounded=0) = Count difference between last and first item in the list",
                    "change_operator(list_of_lists, rounded=0, first_last=False) = This operator combine change functions",
                    "percent_layout(list_of_lists, rounded=1) = Percent layout inside the lists",
                    "percent_layout_in_more_lists (list_of_lists, rounded=1) = Percent layout between the lists in the index possitions.",
                    "percent_change(list_of_values, rounded=1) = Create the list of the percent change",
                    "percent_change_fl(list_of_values, rounded=1) = Create the list of the percent change between the first and the last item",
                    "percent_change_choice(list_of_values, present_index, past_index, rounded=1) = Create the list of the percent change between two choiced numbers (indexes)",
                    "percent_change_space(list_of_values, space=2, rounded=1) = Create the list of the percent change with specified space [space = 4; past = index 0, present = index 4]",
                    "percent_change_skip(list_of_values, skip, rounded=1) = Create the list of the percent change with skip range [skip = 4; past = index -5, present = index -1]",
                    "percent_change_operator(list_of_lists, rounded=1, first_last=False, choice=None, space=None, skip=None) = This operator combine percent_change functions",
                    "percent_change_two_lists = Percent changes between values in two lists",
                    "percent_operation(A, B, A_percent=100, B_percent=False) = Percent operation - find the unknown",
                    "rounder(list_of_lists, rounded=0, floating=True, remove_string=False) = Creating the new lists of the rounded numbers",
                    "modulo(a_num, b_num) = Counting modulo and floor division; a_num devided by b_num",
                    "median(list_of_values) = Median, or 50th percentile, of grouped data",
                    "csv_writer (filename, list_of_labels, list_of_lists) = Creating csv file",
                    "csv_reader(filename, encoding) =  Reading csv file and making the lists",
                    "list_operator (num_list, operation_num, operator, rounded=1) = num_list[i] operator (default = *) operation_num",
                    "list_to_lists(list_of_values, num_of_lists) = Transform one data list to more lists",
                    "lists_to_list(list_of_values) = Nested lists transform to the simple list",
                    "list_int(list_of_values, no_int_remove=False) = Creating list of integers",
                    "list_float(list_of_values, no_float_remove=False) = Creating list of float nums",
                    "list_str(list_of_values) = Creating list of string",
                    "lists_sum(list_of_lists, rounded=2) = Take lists and count sum of each list",
                    "list_sum_division(list_of_values, rounded=1) = Sum of nums divided by len of list",
                    "lists_sum_division(*args, rounded=1) = Take the lists and count average",
                    "list_from_index(list_of_lists, index) = Creating new list from given index in lists",
                    "lists_from_indexes(list_of_lists, index_range=None) = Create lists in list from indexes (columns)",
                    "list_from_indexes_and_sum_division(list_of_lists, index_range=None, rounded=1) = Lists_from_indexes + lists_sum_division",
                    "quartal_list(start, end, space=1) = Creating quartal list",
                    "months_cz_list(start, end, space=1) = Creating months CZ list",
                    "months_num_list(start, end, space=1) = Creating months num list",
                    "for_string_in_range(first_str, second_str, start, end, space=1) = Creating desired string with nums",
                    "test_shape(*args, lol=None, len_num=None) = Testing shape; lol = list of list",
                    "list_to_dict_sort_and_back (list_one, list_two, reverse=False) = Take two lists - create dictionary - sort values - create two list sorted according values",
                    "lists_dict_sort(list_of_lists, reverse=False) = Take lists - create dictionary - SORT BY VALUES - !SORTING LIST = INDEX 1!",
                    "two_lists_difference (List_A, List_B, rounded=1) = Values from List_A minus values from List_B = counting difference",
                    "two_lists_difference_operator(A_list_of_lists, B_lists_of_lists, percent=False, rounded=0) = Using two_lists_difference + percent_change_two_lists",
                    "per_population (values_list, population_list, per_population=1_000, rounded=0, reverse=False) = Values devided per population",
                    "per_population_in_time(list_of_values_lists, list_of_population_lists, per_pop=1_000, rounded=0, reverse=False) = per_population extended to counting in time",
                    "decrease_or_increase_numbers (list_of_lists, operation_number, increase=False, rounded=0) = Numbers in each list increased or decreased by operation_number",
                    "analyse_num_list(num_list) = Analysing num list",
                    "remove_string_from_list(list_of_values) = Remove strings from list",
                    "remove_string_from_lists(list_of_lists) = Remove strings from lists",
                    "remove_letters_from_string(list_of_values) = Take the list and from each string remove letters",
                    "excel_writer(filename, list_of_labels, list_of_lists, write_lables=True) = Write data to xlsx",
                    "excel_reader(filename, sheet_name, list_feed_column=False, if_cell_none, remove_string=False, len_columns=None) = Read data from xlsx",
                    "print_item_index(list_with_items) = Print index of item in the list",
                    "print_table(list_of_lists, reshape=False) = Print data in the lists",
                    "lists_function_operator(list_of_lists, func_name_without_brackets) = Take function for one list and extend it for lists",
                    "lists_function_operator_integrated(many predefined kwargs) = Take function for one list and extend it for lists",
                    "ordered_lists(list_of_lists, reverse=True) = Take lists - create dictionary - SORT BY KEYS - !SORTING LIST = INDEX 0!",
                    "dictator(list_of_lists, reverse=False, sorted_by_keys=False) = Take lists - create dictionary and sort; default sort by values",
                    "cumulation(list_of_lists, rounded=0) = Count cumulation in the list(s)",
                    "remove_items(list_of_items, list_of_removing_items, remove_by_index=False) = Remove desired items from list",
                    "txt_reader (filename, split=True) = Read txt files",
                    "text_analyse(analysed_list, print_items=False, sliced=None, len_pass=1, unwanted=True, unwanted_list=None, lower_case=True) = Analyse list with strings",
                    "edit_num_format(list_of_values, crash_sign) = Edit nums from non transforable format to trasformable format",
                    "data_shortcut (list_of_lists, jump=2, start_from_end=True, rounded=2) = Create smaller dataset from the big one",
                    "find_indexes(list_of_values, keyword, return_all=False) = Search in large dataset for keyword and find out index position",
                    "find_data_in_range(list_of_values, start, end) = Find specific data in range of bigger dataset",
                    "find_indexes_in_range(list_of_values, start, end) = Find indexes in specific range",
                    "find_data_in_range_operator(list_of_lists, keyword, start, end, locators=[1,2,3], rounded=0) = Shortcut operator for searching in the dataset in range",
                    "find_dataset_in_range(list_of_lists, list_of_keywords, start, end, locators=[1,2,3], rounded=0) = Upgraded func find_data_in_range - you can search for list of keywords",
                    "find_labels(list_of_labels) = Find labels in dataset and avoid to duplicate",
                    "without_currency(list_of_values, rounded=1) = Only nums, without currency = 40Kč to 40",
                    "append_currency(list_of_list, currency) = Take int, creates str, put currency and space between each three nums",
                    "find_rows(list_of_lists, list_of_labels, reshape=False, sliced=None,) = Find the row in the dataset; searchind according to list_of_labels",
                    "unpack_list(list) = Unpack packed lists",
                    "reshape(list_of_lists, index_range=None) = Same fn as lists_from_indexes - only renamed for faster using",
                    "print_not_found(list_of_lists, list_of_labels) = Find if label/key is not in the lists",
                    "today_date() = Return today date",
                    "search_html(url, find, attrs={'class':'container'}, selected=None, data_format) = Searching data in the html page",
                    "download_csv_file(url, filename) = Paste url of csv file and it will be downloaded",
                    "class WebCall = Working with urls; using requests module",
                    ]

    list_of_funcs.sort()
    print("\nATools = Analytical Tools")
    print("\n:: THE FUNCTIONS LIST ::\n")
    index = 1
    for function in list_of_funcs:
        if index < 10:
            index = (str(index))
            index = "0" + index
        print(f"{index} --- {function}")
        index = (int(index))
        index +=1
    print(f"\nNumber of functions: {len(list_of_funcs)}\n\nATools version: {ver} ({verdate})\nCreated by {author} ({created})")
    print(f"More than {Lines} lines of code")

    if __name__ == "__main__":
        __ = input("")

def rounder(list_of_lists, rounded=0, floating=True, remove_string=False):
    "Creating the new lists of the rounded numbers"

    data = []
    lol=list_of_lists
    
    for li in lol:
        work=[]

        for n in li:
            try:
                if floating==True:
                    n = (float(n))
                n = (round(n, rounded))
                if rounded==0:
                    n = (int(n))
                work.append(n)

            except ValueError:
                if remove_string==False:
                    work.append(n)
                else:
                    pass

        data.append(work)

    if (len(data)) == 1:
        return data[0]

    else:
        return data 

def print_not_found(list_of_lists, list_of_labels):
    """Find if label/key is not in the lists"""

    data = list_of_labels
    lol = list_of_lists

    for l in lol:
        for lab in data:
            if lab in l:
                data.remove(lab)
                break

    #for label in data:
    #    for li in lol:
    #        if label in li:
    #            data.remove(label)
    #            break

    if (len(data)) != 0:
        num = 1
        for d in data:
            print(f"{num} --- {d} --- NOT FOUND!")
            num += 1

def change(list_of_values, rounded=0):
    """Count change: index 1 - index 0; i 2 - i 1 etc."""

    lov = list_of_values
    data = []
    for i in range (len(lov)):
        if i == 0:
            pass
        else:
            x = lov[i] - lov[i-1]
            data.append(x)

    x = rounder([data], rounded=rounded)

    return x

def change_fl(list_of_values, rounded=0):
    """Count difference between last and first item in the list"""

    data = []

    lov = list_of_values

    x = lov[-1] - lov[0]
    data.append(x)

    result = rounder([data], rounded=rounded)
    return result

def change_operator(list_of_lists, rounded=0, first_last=False):
    """This operator combine change functions;
    change: Count change: index 1 - index 0; i 2 - i 1 etc.
    change_fl: Count difference between last and first item in the list"""

    data = []

    for li in list_of_lists:
        if first_last != False:
            x = change_fl(li, rounded=rounded)

        else:
            x = change(li, rounded=rounded)

        data.append(x)
    if (len(data)) == 1:
        return data[0]
    else:
        return data

def percent_layout(list_of_lists, rounded=1):
    """Percent layout inside the lists"""

    data=[]
    lol=list_of_lists

    for li in lol:
        full = (sum(li))
        work=[]
        for n in li:
            x = n / full
            x = x*100
            work.append(x)
        data.append(work)

    r = rounder(data, rounded=rounded)
    return r

def percent_layout_in_more_lists (list_of_lists, rounded=1):
    """Percent layout between the lists
        in the index possitions. Percent layout 
        between indexes 1, between indexes 2 etc. """

    lol = list_of_lists

    final = []
    for l in lol: final.append([])

    for i in range(len(lol[0])):

        full_num = 0

        for l in lol:
            full_num += l[i]

        list_index = 0
        for l in lol:
            percent = (l[i] / full_num) * 100
            percent = (round(percent, rounded))
            final[list_index].append(percent)
            list_index += 1

    return final

def percent_change(list_of_values, rounded=1):
    """Create the list of the percent change
        Growth Rate = ((present - past) / past) * 100 """

    change, working_list = [], []
    working_list.append(list_of_values[0])

    for num in list_of_values[1:]:
        working_list.append(num)
        growth_rate = (((working_list[-1] - working_list[-2]) / working_list[-2]) * 100 )
        growth_rate = (round(growth_rate, rounded))
        if rounded == 0:
            growth_rate = (int(growth_rate))
        change.append(growth_rate)

    return change

def percent_change_fl(list_of_values, rounded=1):
    """Create the list of the percent change between the first and the last item
        Growth Rate = ((present - past) / past) * 100 """

    change = []

    growth_rate = (((list_of_values[-1] - list_of_values[0]) / list_of_values[0]) * 100 )
    growth_rate = (round(growth_rate, rounded))
    if rounded == 0:
        growth_rate = (int(growth_rate))
    change.append(growth_rate)

    return change

def percent_change_choice(list_of_values, present_index, past_index, rounded=1):
    """Create the list of the percent change between two choiced numbers / indexes
        Put indexes of the numbers to the parameters
        Growth Rate = ((present - past) / past) * 100 """

    change = []

    growth_rate = (((list_of_values[present_index] - list_of_values[past_index]) / list_of_values[past_index]) * 100 )
    growth_rate = (round(growth_rate, rounded))
    if rounded == 0:
        growth_rate = (int(growth_rate))
    change.append(growth_rate)

    return change

def percent_change_space(list_of_values, space=2, rounded=1):
    """Create the list of the percent change with specified space [space = 4; past = index 0, present = index 4]
        Growth Rate = ((present - past) / past) * 100 """

    change = []
    present_index = (len(list_of_values)) - 1
    past_index = present_index - space
    
    while past_index >= 0:
        growth_rate = (((list_of_values[present_index] - list_of_values[past_index]) / list_of_values[past_index]) * 100 )
        growth_rate = (round(growth_rate, rounded))
        if rounded == 0:
            growth_rate = (int(growth_rate))
        change.insert(0, growth_rate)
        present_index -= 1 
        past_index -= 1

    return change

def percent_change_skip(list_of_values, skip, rounded=1):
    """Create the list of the percent change with skip range [skip = 4; past = index -5, present = index -1]
        Growth Rate = ((present - past) / past) * 100 """

    change = []
    value = skip
    cursor = skip + 1
    present = -1
    past = -cursor

    while cursor < (len(list_of_values)):
        growth_rate = (((list_of_values[present] - list_of_values[past]) / list_of_values[past]) * 100 )
        growth_rate = (round(growth_rate, rounded))
        if rounded == 0:
            growth_rate = (int(growth_rate))
        change.insert(0, growth_rate)
        cursor += value
        present -= value
        past -= value 

    return change

def percent_change_two_lists (past_list, present_list, rounded=1):
    """Percent changes between values in two lists
    Growth Rate = ((present - past) / past) * 100 """

    data = []
    len_data = (len(past_list))

    for i in range (len_data):
        x = ((present_list[i] - past_list[i]) / past_list[i]) * 100
        x = round(x, rounded)
        if rounded == 0:
            x = (int(x))
        data.append(x)

    return data

def percent_change_operator(list_of_lists, rounded=1, first_last=False, choice=None, space=None, skip=None):
    """This operator combine percent_change functions;
    percent_change: Create the list of the percent change;
    percent_change_fl: Create the list of the percent change between the first and the last item;
    percent_change_choice: Create the list of the percent change between two choiced numbers / indexes (choice=[past_index,present_index]);
    percent_change_space: Create the list of the percent change with specified space [space = 4; past = index 0, present = index 4];
    percent_change_skip: Create the list of the percent change with skip range [skip = 4; past = index -5, present = index -1]
        Growth Rate = ((present - past) / past) * 100"""

    data = []

    for li in list_of_lists:
        if first_last != False:
            x = percent_change_fl(li, rounded=2)
        
        elif choice:
            x = percent_change_choice(li, choice[1], choice[0], rounded=2)

        elif skip:
            x = percent_change_skip(li, skip, rounded=2)

        elif space:
            x = percent_change_space(li, space=space, rounded=2)

        else:
            x = percent_change(li, rounded=2)

        data.append(x)

    result = rounder(data, rounded=rounded)
    return result

def percent_operation(A, B, A_percent=100, B_percent=False):
    """Percent operation - find the unknown"""

    if B_percent==False:
        x = A / A_percent
        y = B / x
        y = (round(y, 2))
        difference = A_percent - y
        difference = (round(difference, 2))
        print(f"{A} is {A_percent} %\n{B} is {y} %\nDifference is {difference} %")

    else:
        x = A / A_percent
        y = x * B
        y = (round(y, 2))
        difference = A_percent - B
        difference = (round(difference, 2))
        print(f"{A} is {A_percent} %\n{y} is {B} %\nDifference is {difference} %")

def modulo(a_num, b_num):
    """
    Counting modulo and floor division; a_num devided by b_num
    """  

    division = a_num / b_num
    floor_division = a_num // b_num
    floor_division = int(floor_division)
    modulo = a_num % b_num
    
    print(a_num, " devided by ", b_num, " is: ", division)
    print("Floor division / Without remain: ", floor_division)
    print("Remain is / Modulo: ", modulo)

def csv_writer (filename, list_of_labels, list_of_lists, header=True, encoding="utf8"):
    """Creating csv file; using csv module
    header = print head row with labels / print labels"""

    import csv
  
    with open(filename, mode='w', newline="", encoding=encoding) as csv_file:
        file_writer = csv.writer(csv_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        #OR: file_writer = csv.writer(csv_file)

        if header == True:
            file_writer.writerow(list_of_labels)

        len_data = (len(list_of_lists[0]))
        index = len_data 
        row_range = (len(list_of_lists)) 

        for i in range (index):
            row = []
            for r in range (row_range):
                try:
                    x = list_of_lists[r][i]
                    row.append(x)
                except IndexError:
                    x = None
                    row.append(x)

            file_writer.writerow(row)

def csv_reader(filename, encoding="utf8"):
    """Reading csv file and making the lists
    using csv module"""

    import csv

    data = []
  
    with open(filename, newline='', encoding=encoding) as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if (len(row)) == 0:
                pass
            
            else:
                x = row
                data.append(x)
                                  
    return data

def list_to_lists(list_of_values, num_of_lists):
    """Transform one data list to more lists"""
    
    lol = []
    for n in range(num_of_lists):
        n = []
        lol.append(n)

    space = num_of_lists
    index_cursor = space - 1

    while index_cursor >= 0:
        
        cursor = 1
        loop_cursor = cursor
        current_index = index_cursor
        current_list = index_cursor

        while loop_cursor <= (len(list_of_values)) / num_of_lists:
            x = list_of_values[current_index]
            lol[current_list].append(x)
            current_index += space
            loop_cursor += 1

        index_cursor -= 1

    #reverse lol list
    return lol

def lists_to_list(list_of_values):
    """Nested lists transform to the simple list"""

    data = []
    for nested in list_of_lists:
        for n in nested:
            data.append(n)
    return data

def median(list_of_values):
    """Median, or 50th percentile, of grouped data
    using statistics module"""

    import statistics

    median = statistics.median_grouped(list_of_values)

    return median

def list_operator (num_list, operation_num, operator="*", rounded=1):
    """num_list[i] operator (default = *) operation_num"""

    data = []
    len_data = (len(num_list))

    for i in range (len_data):
        if operator == "+":
            x = num_list[i] + operation_num
        elif operator == "-":
            x = num_list[i] - operation_num
        elif operator == "/":
            x = num_list[i] / operation_num
        elif operator == "//":
            x = num_list[i] // operation_num
        elif operator == "%":
            x = num_list[i] % operation_num
        else:
            x = num_list[i] * operation_num
        x = (round(x, rounded))
        if rounded == 0:
            x = (int(x))
        data.append(x)

    return data

def list_int(list_of_values, no_int_remove=False):
    """Creating list of integers
    Float nums will be rounded"""

    values = []
    for l in list_of_values:
        try:
            l = (float(l))
            l = (round(l))
            l = (int(l))
            values.append(l)

        except ValueError:
            if no_int_remove == False:
                values.append(l)
            else:
                pass

    return values

def list_float(list_of_values, no_float_remove=False):
    """Creating list of float nums"""

    values = []
    for l in list_of_values:
        try:
            l = (float(l))
            values.append(l)
        except ValueError:
            if no_float_remove == False:
                values.append(l)
            else:
                pass

    return values

def list_str(list_of_values):
    """Creating list of string"""

    values = []
    for l in list_of_values:
        l = (str(l))
        
        values.append(l)

    return values

def lists_sum(list_of_lists, rounded=2):
    """Take lists and count sum of each list"""

    data = []

    for li in list_of_lists:
        li = sum(li)
        data.append(li)

    x = rounder([data], rounded=rounded)
    return x

def list_sum_division(list_of_values, rounded=1):
    """Sum of nums divided by len of list"""

    sum_nums = (sum(list_of_values))
    len_list = (len(list_of_values))
    x = sum_nums / len_list
    x = (round(x,rounded))
    if rounded == 0:
        x = (int(x))

    return x

def lists_sum_division(*args, rounded=1):
    """Take the lists and count average"""

    li = []
    for l in args:
        x = (len(l))
        y = (sum(l))
        z = y / x
        z = (round(z, rounded))
        if rounded == 0:
            z = (int(z))
        li.append(z)

    return li

def list_from_index(list_of_lists, index):
    """Creating new list from given index in lists"""

    index_list = []
    for l in list_of_lists:
        l = l[index]
        index_list.append(l)

    return index_list

def lists_from_indexes(list_of_lists, index_range=None):
    """Create lists in list from indexes (columns)"""

    index_list = []

    if index_range != None:
        index_range = index_range + 1
    else:
        index_range = (len(list_of_lists[0]))

    for n in range (index_range):
        li = []
        for l in list_of_lists:
            try:
                l = l[n]
                li.append(l)

            except IndexError:
                li.append(None)

        index_list.append(li)

    return index_list

def list_from_indexes_and_sum_division(list_of_lists, index_range=None, rounded=1):
    """Lists_from_indexes + lists_sum_division"""

    a = lists_from_indexes(list_of_lists, index_range=index_range)
    b = []
    c = (len(a))
    for n in range (c):
        x = (len(a[n]))
        y = (sum(a[n]))
        z = y / x
        z = (round(z, rounded))
        if rounded == 0:
            z = (int(z))
        b.append(z)

    return b

def quartal_list(start, end, space=1):
    """Creating quartal list"""

    q_list = []

    for n in range (start, end+1, space):
        n = (str(n))
        q1 = "Q1 " + n
        q2 = "Q2 " + n
        q3 = "Q3 " + n
        q4 = "Q4 " + n
        q_list.append(q1)
        q_list.append(q2)
        q_list.append(q3)
        q_list.append(q4)

    return q_list

def months_cz_list(start, end, space=1):
    "Creating months CZ list"

    cz_list = []

    for n in range(start, end +1, space):
        n = (str(n))
        m1 = "Leden " + n; cz_list.append(m1)
        m2 = "Únor " + n; cz_list.append(m2)
        m3 = "Březen " + n; cz_list.append(m3)
        m4 = "Duben " + n; cz_list.append(m4)
        m5 = "Květen " + n; cz_list.append(m5)
        m6 = "Červen " + n; cz_list.append(m6)
        m7 = "Červenec " + n; cz_list.append(m7)
        m8 = "Srpen " + n; cz_list.append(m8)
        m9 = "Září " + n; cz_list.append(m9)
        m10 = "Říjen " + n; cz_list.append(m10)
        m11 = "Listopad " + n; cz_list.append(m11)
        m12 = "Prosinec " + n; cz_list.append(m12)

    return cz_list

def months_num_list(start, end, space=1):
    "Creating months num list"

    num_list = []

    for n in range(start, end +1, space):
        n = (str(n))
        m1 = "01/" + n; num_list.append(m1)
        m2 = "02/" + n; num_list.append(m2)
        m3 = "03/" + n; num_list.append(m3)
        m4 = "04/" + n; num_list.append(m4)
        m5 = "05/" + n; num_list.append(m5)
        m6 = "06/" + n; num_list.append(m6)
        m7 = "07/" + n;num_list.append(m7)
        m8 = "08/" + n; num_list.append(m8)
        m9 = "09/" + n; num_list.append(m9)
        m10 = "10/" + n; num_list.append(m10)
        m11 = "11/" + n; num_list.append(m11)
        m12 = "12/" + n; num_list.append(m12)

    return num_list

def for_string_in_range(first_str, second_str, start, end, space=1):
    "Creating desired string with nums"

    str_list = []

    for n in range (start, end + 1, space):
        n = (str(n))
        x = first_str + n + second_str
        str_list.append(x)

    return str_list

def test_shape(*args, lol=None, len_num=None):
    """Testing shape; lol = list of list"""

    li = []

    if lol:
        for l in lol:
            l = (len(l))
            li.append(l)

    else:
        if len_num:
            for l in args:
                len_list = (len(l))
                if len_list == len_num:
                    li.append(len_list)

                else: 
                    li.append(len_list)
                    err = (len(l))
                    print(f"{l} = {err}")

        else:
            for l in args:
                l = (len(l))
                li.append(l)

    print(li)

def list_to_dict_sort_and_back (list_one, list_two, reverse=False):
    """Take two lists - create dictionary - sort values - create two list sorted according values"""

    result = [[], []]

    lists_to_dict = dict(zip(list_one, list_two))
    ltd = lists_to_dict

    if reverse != False:
        sorted_dict = sorted(ltd.items(), key = lambda x: x[1], reverse = False)
        #sorted_dict = dict(sorted(ltd.items(), key=lambda item: item[1]))
        #sorted_dict = {k: v for k, v in sorted(ltd.items(), key=lambda item: item[1])}
    else:
        sorted_dict = sorted(ltd.items(), key = lambda x: x[1], reverse = True)

    for sd in sorted_dict:
        key = sd[0]
        value = sd[1]
        result[0].append(key)
        result[1].append(value)

    return result

def lists_dict_sort(list_of_lists, reverse=False):
    """Take lists - create dictionary - SORT BY VALUES - !SORTING LIST = INDEX 1!
    using func: list_to_dict_sort_and_back"""

    data=[]
    
    lol = list_of_lists
    lel = (len(list_of_lists))
    for i in range (lel):
        index = i
        if index == 1:
            index = 0
        x = list_to_dict_sort_and_back(lol[index], lol[1], reverse=reverse)
        if i == 1:
            data.append(x[1])
        else:
            data.append(x[0])

    return data

def two_lists_difference (List_A, List_B, rounded=1):
    """Values from List_A minus values from List_B = counting difference"""

    data = []
    len_data = (len(List_A))

    for i in range (len_data):
        x = List_A[i] - List_B[i]
        x = round(x, rounded)
        if rounded == 0:
            x = (int(x))
        data.append(x)

    return data

def per_population (values_list, population_list, per_population=1_000, rounded=0, reverse=False):
    """Values devided per population
    reverse = reverse operation; You have data per_pop and you will get absolute numbers"""

    data = []
    len_data = (len(values_list))

    for i in range (len_data):
        if reverse==False:
            x = population_list[i] / per_population
            y = values_list[i] / x 
        else:
            #x = population_list[i] * per_population
            #y = values_list[i] * x 
            x = values_list[i] * population_list[i]
            y = x / per_population
        y = round(y, rounded)
        if rounded == 0:
            y = (int(y))

        data.append(y)

    return data

def per_population_in_time(list_of_values_lists, list_of_population_lists, per_pop=1_000, rounded=0, reverse=False):
    """per_population (values devided by population) extended to counting in time;
    list_of_values_lists[0] devided by list_of_population_lists[0];
    per_pop = per_population
    reverse = reverse operation; You have data per_pop and you will get absolute numbers"""

    data=[]
    v = list_of_values_lists
    p = list_of_population_lists
    for i in range (len(v)):
        if (len(v[i])) != (len(p[i])):
            x = (len(v[i]))
            y = (len(p[i]))
            return ["ERR SHAPE", x, y]
        else:
            x = per_population(v[i], p[i], per_population=per_pop, rounded=rounded, reverse=reverse)
        data.append(x)

    if (len(data)) == 1:
        return data[0]
    else:
        return data

def decrease_or_increase_numbers(list_of_lists, operation_number, increase=False, rounded=0):
    """Numbers in each list increased or decreased by operation_number;
    Division or multiplication"""

    data = []
    lol = list_of_lists

    for li in lol:
        work = []
        for i in li:

            if increase != False:
                y = i * operation_number  
            else:
                y = i / operation_number 
            work.append(y)

        data.append(work)

    result = rounder(data, rounded=rounded)

    return result

def analyse_num_list(num_list):
    """Analysing num list"""

    print("-----------------\nANALYSE NUM LIST:\n-----------------")
    a = (len(num_list)); a = (round(a, 2))
    print(f"Len = {a}")
    b = (sum(num_list)); b = round(b, 2)
    print(f"Sum = {b}")
    c = list_sum_division(num_list, rounded=2)
    print(f"Avg = {c}")
    d = median(num_list)
    print(f"Median = {d}")
    print("-----------------")
    sorted_list = (sorted(num_list))
    percent_operation(sorted_list[-1], sorted_list[0])

def remove_string_from_list(list_of_values):
    """Remove strings from list"""

    data = []

    for i in list_of_values:
            
        try:
            x = (float(i))
            data.append(i)
        except ValueError:
            pass
    
    return data

def remove_string_from_lists(list_of_lists):
    """Remove strings from lists"""

    data = []

    for li in list_of_lists:
        x = remove_string_from_list(li)
        data.append(x)

    return data

def remove_letters_from_string(list_of_values):
    "Take the list and from each string remove letters"

    new = []

    for word in list_of_values:
        word = (str(word))
        new_word = ""

        for letter in word:
            try: 
                n = (int(letter))
                n = (str(letter))
                new_word += n

            except ValueError:
                pass
        
        new.append(new_word)
        new = list_int(new)

    return new 

def excel_writer(filename, list_of_labels, list_of_lists, write_lables=True, len_data=None):
    """Write data to xlsx
    using openpyxl library"""
    from openpyxl import Workbook

    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Data can be assigned directly to cells
    #ws['A1'] = 42

    # Rows can also be appended
    if write_lables==True:
        ws.append(list_of_labels)

    lengh = (len(list_of_lists[0]))
    index = lengh
    if len_data:
        index = len_data 
    row_range = (len(list_of_lists)) 

    for i in range (index):
        row = []
        for r in range (row_range):
            try:
                x = list_of_lists[r][i]
                row.append(x)
            except IndexError:
                x = None
                row.append(x)

        ws.append(row)

    # Save the file
    wb.save(filename)

def excel_reader(filename, sheet_name="List1", list_feed_column=False, if_cell_none="empty", remove_string=False, len_columns=None):
    """Read data from xlsx
    if_cell_none = "pass" => pass; if_cell_none = "None" => append None
    using openpyxl library"""

    from openpyxl import load_workbook

    wb = load_workbook(filename)
    ws = wb[sheet_name]

    data = []
          
    row_index = 0
    for row in ws.rows:
        r = []
        data.append(r)
        for cell in row:
            #print(cell.value)
            cv = cell.value
            if cv == None:
                if if_cell_none != "pass" and "None":
                    cv = if_cell_none
                    data[row_index].append(cv)

                elif if_cell_none == "None":
                    cv = None
                    data[row_index].append(cv)

                else:
                    pass

            else:
                data[row_index].append(cv)
        row_index += 1
    
    data_two = []
    if list_feed_column != False:
        len_data = (len(data))

        len_of_columns = (len(data[0]))
        if len_columns != None:
            len_of_columns = len_columns

        for l in range (len_of_columns):
            l = []
            data_two.append(l)

        for li in range (len_data):
                for i in range (len_of_columns):
                    try:
                        data_two[i].append(data[li][i])
                    except IndexError:
                        pass

    if remove_string == True and list_feed_column != False:
        remove_string_from_lists(data_two)
    if remove_string == True and list_feed_column == False:
        remove_string_from_lists(data)

    wb.close()
    if list_feed_column != False:
        return data_two
    else:
        return data

def print_item_index(list_with_items):
    """Print index of item in the list"""

    print("--------------\nindex --- item\n--------------")
    index = 0
    for i in list_with_items:
        print(f"{index} --- {i}")
        index += 1

def lists_function_operator(list_of_lists, func_name_without_brackets):
    """Take function for one list and extend it for lists
    Working only with func with one required parameter"""

    data = []

    for li in list_of_lists:        
        y = func_name_without_brackets(li)    
        data.append(y)

    return data

def ordered_lists(list_of_lists, reverse=True):
    """Take lists - create dictionary - SORT BY KEYS - !SORTING LIST = INDEX 0!
    using collections module"""
    from collections import OrderedDict

    data = []
    
    for li in list_of_lists:
        li=[]
        data.append(li)

    lol = list_of_lists
    lel = (len(lol))

    for i in range(lel):
        
        x = (dict(zip(lol[0], lol[i])))
        ordered = OrderedDict(sorted(x.items()))
        val = ordered.values()
        #data.append(val)
        
        for v in val:
            data[i].append(v)

    if reverse == True:
        pass
        
    else:
        for i in range(lel):
            data[i].reverse()

    return data  

def dictator(list_of_lists, reverse=False, sorted_by_keys=False):
    """Take lists - create dictionary and sort; default sort by values;
        if sort by values = sorting list index 1;
        elif sort by keys = sorting list index 0;
        using func lists_dict_sort and ordered_lists"""

    if sorted_by_keys==False:
        by_values = lists_dict_sort(list_of_lists, reverse=reverse)
        return by_values

    else:
        by_keys = ordered_lists(list_of_lists, reverse=reverse)
        return by_keys

def print_table(list_of_lists, reshape=False):
    """Print data in the lists
    reshape = lists_from_indexes"""

    lol = list_of_lists
    if reshape != False:
        lol = lists_from_indexes(lol)

    lel = (len(lol[0]))
    lelli = (len(lol))
    index = 0

    print("----------------\nINDEX --- ITEMS\n----------------")
    for i in range (lel):
        work = []
        for li in range (lelli):
            try:
                x = lol[li][i]
                work.append(x)
            except IndexError:
                x = None
                work.append(x)
        print(f"{index} --- {work}")
        index += 1
    
def cumulation(list_of_lists, rounded=0):
    """Count cumulation in the list(s);
    index 0 + index 1 = a; index 2 + a = b; index 3 + b = c""" 

    data = []
    lol = list_of_lists
    lel = (len(lol))

    for li in range (lel):
        work = []
        for i in range (len(lol[li])):
            if i == 0:
                work.append(lol[li][0])
            else:
                x = lol[li][i] + work[i - 1]
                work.append(x)
        work_rounded = []

        for w in work:
            w = round(w, rounded)
            if rounded==0:
                w = (int(w))
            work_rounded.append(w)

        data.append(work_rounded)

    one_list = []
    if lel == 1:
        for n in data[0]:
            one_list.append(n)
        return one_list
    else:
        return data

def remove_items(list_of_items, list_of_removing_items, remove_by_index=False):
    """Remove desired items from list"""

    lori = list_of_removing_items
    loi = list_of_items
    loi = loi[:]

    if remove_by_index==False:
        for item in lori:
            loi.remove(item)

    else:
        lori=sorted(lori)
        lori.reverse()
        for item in lori:
            del loi[item]

    return loi

def txt_reader (filename, split=True, encoding="utf-8"):
    """Read txt files"""
    
    with open(filename, encoding=encoding) as file_object:
        
        content = file_object.read()

        if split == True:
            content = content.split()
        else:
            pass

    return content

def text_analyse(analysed_list, print_items=False, sliced=None, len_pass=1, unwanted=True, unwanted_list=None, lower_case=True):
    """Analyse list with strings - function with many parameters;
    print_items = result clearly printed;
    sliced = from long list get shorter;
    len_pass = how long words will be automatically passed;
    unwanted = unwanted list will pass predefined words;
    unwanted_list = user can append next unwanted words;
    lower_case = each word is counted with lower case"""

    unwanted_words = ["pro", "do", "pak", "ke", "od", "na", "se", "je", "si", "ale", "za", "co",
                        "tak", "jak", "ve", "by", "jako", "aby", "nebo", "jsou", "po", "kde", "jen", "taky",
                        "který", "která", "které", "kteří", "když", "mají", "ještě",  
                        "the", "for", "from", "and", "of", "in", "on", "to",
                        "že", "má", "případně", "při", "bez", "také", "tyto", "může", "tedy", "těchto", "nejen",
                        "tomu", "pouze", "ze", "až", "být", "další", "případně", "ani", "tento", "tato", "tzv", "proto", 
                        "dále", "však", "zcela",
                        "tím", "již", "toho", ]
    if unwanted_list:
        for word in unwanted_list:
            unwanted_words.append(word)

    if unwanted != True:
        unwanted_words = []

    keys = []
    values = []
    for word in analysed_list:

        def test_word (word):
            test = ""
            unwanted = [".", ",", ":", ";", "!", "?"]
            for l in word:
                if l in unwanted:
                    pass
                else:
                    test += l
            return test
        
        word = test_word(word)

        try:
            (float(word))
            continue
        except ValueError:
            pass

        if lower_case==True:
            word = word.lower()
        else:
            pass

        if (len(word)) <= len_pass or word in unwanted_words:
                pass
        else:
            if word not in keys:
                keys.append(word)
                values.append(1)
            else:
                index = keys.index(word)
                values[index] += 1

    x = dictator([keys, values])

    if print_items != False:
        place = 1
        print(f"--------------------\nPLACE | WORD | USED\n--------------------")
        for n in range (100):
            
            try:
                print(f"{place} --- {x[0][n]} --- {x[1][n]}x")
            except IndexError:
                pass
            place += 1

    if sliced:
        x[0] = x[0][:sliced]
        x[1] = x[1][:sliced]

    return x

def edit_num_format(list_of_values, crash_sign=" ", replace_with_dot=False):
    """Edit nums from non transforable format to trasformable format;
    crash_sign is a character which do not allow python to recognize the number (space, slash, etc.)
    basic replace sign is "_" or you can choose replace_with_dot"""

    data=[]
    for string in list_of_values:
        number =""
        for char in string:
            if char == crash_sign:
                if replace_with_dot == False:
                    char = "_"
                else:
                    char = "."

                number += char
            else:
                number += char

        data.append(number)
    return data    

def data_shortcut (list_of_lists, jump=2, start_from_end=True, rounded=2):
    """Create smaller dataset from the big one;
    jump = new list will leave only every second (or pick different) index"""

    data = []

    for li in list_of_lists:
        index = 0
        work = []

        if start_from_end == True:
            work_data = li[:: -1] # reverse list
        else:
            work_data = li

        while index < (len(work_data)):
            x = work_data[index]
            work.append(x)
            index += jump
        
        if start_from_end == True:
            work = work[:: -1]

        data.append(work)

    x = rounder(data, rounded=rounded)

    return x

def two_lists_difference_operator(A_list_of_lists, B_lists_of_lists, percent=False, rounded=0):
    """This operator works with bigger datasets and using two_lists_difference + percent_change_two_lists;
    Find out difference between two lists with the same index (A[0] diff B[0]);
    Choose if difference will be in percent or not"""

    data = []

    a = A_list_of_lists
    b = B_lists_of_lists

    lenght = (len(a))

    if percent == False:

        for i in range (lenght):
            x = two_lists_difference(a[i], b[i], rounded=rounded)
            data.append(x)

    else:

        for i in range (lenght):
            x = percent_change_two_lists(a[i], b[i], rounded=rounded)
            data.append(x)   

    if (len(data)) == 1:
        return data[1]

    else:
        return data     

def find_indexes(list_of_values, keyword, return_all=False):
    """Search in large dataset for keyword and find out index position"""

    data = []
    lov = list_of_values

    for i in range(len(lov)):
        if lov[i] == keyword:
            data.append(i)

        else:
            pass
    
    try:
        if return_all==False:
            return (f"{data[0]}:{data[-1] + 1}")
        else:
            return data
    except IndexError:
        return "NOT FOUND"

def find_data_in_range(list_of_values, start, end):
    """Find specific data in range of bigger dataset"""

    data = []
    lov = list_of_values
    lov = list_int(lov) 

    for i in range(len(lov)):
        if lov[i] >= start and lov[i] <= end:
            data.append(lov[i])
        else:
            pass

    return data

def find_indexes_in_range(list_of_values, start, end):
    """Find indexes in specific range"""

    data = []
    lov = list_of_values
    lov = list_int(lov)

    #for i in range(len(lov)):
    #   if lov[i] == start or end:
    #       data.append(i)
    #   else:
    #       pass

    start_index = lov.index(start)
    end_index = lov.index(end)
    return [start_index, end_index + 1]

    #return [data[0], data[-1] + 1]

def find_data_in_range_operator(list_of_lists, keyword, start, end, locators=[1,2,3], rounded=0):
    """Shortcut operator for searching in the dataset in range;
    keyword = looking for specific keyword;
    if keyword == 0: searching only according range (start, end)
    start = where data - which you want - should start (typically years)
    end = where data should end;
    if start + end == 0: data without range - searching only according keyword 
    ! locators=[1,2,3] = keyword will be searched in list 1, start and end point will be searched in list 2 and return data will be taken from list 3;
    This operator using functions - find_indexes + find_indexes_in_range"""

    data = []

    lol = list_of_lists

    if keyword == 0:
        x = [0, 101961]

    else:

        x = find_indexes(lol[locators[0]], keyword, return_all=True)
    
    if (len(x)) == 0:
        return (f"{keyword} IS NOT IN THE LIST")

    else:

        if start + end == 0:

            result = lol[locators[2]][x[0]:x[-1] + 1]

            result = rounder([result], rounded=rounded)

            return result

        else:

            x[-1] += 1

            searched = lol[locators[1]][x[0]:x[-1]]

            y = find_indexes_in_range(searched, start, end)

            start_point = x[0] + y[0]
            end_point = x[0] + y[-1]

            result = lol[locators[2]][start_point:end_point]

            result = rounder([result], rounded=rounded)

            return result

def find_dataset_in_range(list_of_lists, list_of_keywords, start, end, locators=[1,2,3], rounded=0):
    """Upgraded func find_data_in_range - you can search for list of keywords;
    Shortcut operator for searching in the dataset in range;
    keyword = looking for specific keyword;
    if keyword == [0]: searching only according range (start, end)
    start = where data - which you want - should start (typically years)
    end = where data should end;
    if start + end == 0: data without range - searching only according keyword 
    ! locators=[1,2,3] = keyword will be searched in list 1, start and end point will be searched in list 2 and return data will be taken from list 3;
    This operator using functions - find_indexes + find_indexes_in_range"""

    data = []

    for keyword in list_of_keywords:
        x = find_data_in_range_operator(list_of_lists, keyword, start, end, locators=locators, rounded=rounded)
        data.append(x)

    return data

def find_labels(list_of_labels):
    """Find labels in dataset and avoid to duplicate"""

    data = []

    for label in list_of_labels:
        if label in data:
            pass
        else:
            data.append(label)

    return data

def without_currency(list_of_values, rounded=1):
    """Only nums, without currency = 35Kč to 35;
    make str to float"""

    data = []
    for string in list_of_values:
        work = ""
        for s in string:
            if s == ".":
                work += s
            else:
                try:
                    x = (int(s))
                    work += s
                except ValueError:
                    pass

        work = (float(work))
        data.append(work)
    
    result = rounder([data], rounded=rounded)

    return result 

def append_currency(list_of_list, currency=" Kč"):
    """Take int, creates str, put currency and space between each three nums"""

    data = []

    for li in list_of_list:
        nums = []
        for num in li:
            n = ""
            num = (str(num))
            space = 0
            for nn in num[::-1]:
                n += nn
                space +=1
                if space == 3:
                    n += " "
                    space = 0
            n = n[::-1]
            #n += " "
            n += currency
            n = n.lstrip()
            nums.append(n)
        data.append(nums)
    return data

def find_rows(list_of_lists, list_of_labels, reshape=False, sliced=None, not_found=False, integers=True, only_one_row_for_label=True):
    """Find the row in the dataset; searchind according to list_of_labels
        reshape dataset = reverse list, using fn lists_from_indexes
        integers = transform strings to nums if its possible
        sliced = you can choose one index [0], or slice [1:5] - ! type as a list = [1,5] !
        not found = print not found labels 
        only_one_row_for_label = if more rows for one label, it will be ignored"""

    dataset = list_of_lists
    data = []

    if reshape != False:
        dataset = lists_from_indexes(dataset)

    for label in list_of_labels:
        for row in dataset:
            if label in row:
                if row not in data:
                    if label not in data:
                        data.append(row)
                        if only_one_row_for_label == True:
                            break

    if integers == True:
        data_integers = []
        for d in data:
            x = list_int(d)
            data_integers.append(d)
        data = data_integers

    if sliced:
        index = sliced

        data_sliced = []

        if (len(index)) == 1:
            for row in data:
                x = d[index[0]]
                data_sliced.append(x)

        elif (len(index)) == 2:
            for row in data:
                x = d[index[0]:index[1]]
                data_sliced.append(x)

        data = data_sliced               

    if print_not_found != False:
        print_not_found(data, list_of_labels)

    return data

def unpack_list(list):
    """Unpack packed lists"""
    
    return list[0]

def reshape(list_of_lists, index_range=None):
    """Same fn as lists_from_indexes - only renamed for faster using"""

    data = lists_from_indexes(list_of_lists=list_of_lists, index_range=index_range)

    return data

def today_date():
    """Return today date
    using time module"""

    import time

    data = time.strftime("%d. %m. %Y")
    
    return data

def search_html(url, find="div", attrs={'class':'container'}, selected=None, data_format="html"):
    """Searching data in the html page
    url = paste url page

    find and attrs examples = html.find("div", {'class':'container'});
                html.find("table", attrs={"class": "table table-hover"})

    selected example = text.find_all(selected)

    data_format = html or text

    inspect source code of html page 
    
    ver.: 0.2
    """

    import requests
    from bs4 import BeautifulSoup

    #html_content = requests.get(url).text
    #soup = BeautifulSoup(html_content)
    #data = soup.find("table", attrs={"class": "table table-hover"}) 
    #selected = data.find_all("td")

    data = []

    page = requests.get(url)
    html = BeautifulSoup(page.content, "html.parser")

    try:
        if data_format == "html":
            search = html.find(find, attrs=attrs)
        else:
            search = html.find(find, attrs=attrs).text
    except AttributeError:
        pass

    if selected:
        try:
            spec = search.find_all(selected)
            return spec
        except AttributeError:
            return search
    else:
        return search

class WebCall:
    """working with urls; using requests module"""

    def __init__ (self, url, params=None, auth=None, headers=None):
        """basic params about url; payload has to be in dict
        headers example: headers = {"User-Agent": "PostmanRuntime/7.28.0",}"""
        import requests

        try:
            r = requests.get(url, params=params, auth=auth, headers=headers)
            self.r = r
        except ConnectionError:
            print("CONNECTION ERROR")

    def info(self):
        print("---------------\nWebCall - info\n---------------")
        print("URL: ", self.r.url)
        print("ENCODING: ", self.r.encoding)
        print("HEADERS: \n", self.r.headers)
        print("\nTEXT: ", self.r.text)
        print("CONTENT: ", self.r.content)
        print("JSON: ", self.r.json)
        print("---------------")
        print("STATUS: ", self.r.status_code)
        print("---------------")
        
    def status(self):
        print(self.r.status_code)

    def text(self):
        print(self.r.text)

    def json(self):
        print(self.r.json())

def download_csv_file(url, filename):
    """Paste url of csv file and it will be downloaded
    filename = type filename without .csv
    filepath = you can type filepath to the filename param
    code from: kite.com"""

    import requests

    req = requests.get(url)
    url_content = req.content
    csv_file = open(filename, 'wb')

    csv_file.write(url_content)
    csv_file.close()




























def lists_function_operator_integrated(
    list_of_lists, analyse=False, decrease_num=False, increase_num=False, operation_number=None, rounded=0,
    float=False, integer=False, string=False, remove_string=False, list_operator_fn=False, operator="*", rounded_numbers=False):
    """Take function for one list and extend it for lists
    Funcs are integrated"""

    data = []

    if remove_string!=False:
        for li in list_of_lists:
            y = remove_string_from_list(li)
            data.append(y)

    elif analyse!=False:
        index = 0
        for li in list_of_lists:
            print("__________________")
            print("__________________")
            print(f"INDEX --- {index}")
            y = analyse_num_list(li)
            index +=1

    elif decrease_num!=False or increase_num!=False:
        for li in list_of_lists:
            if decrease_num!=False:
                y = decrease_or_increase_numbers(li, operation_number, rounded=rounded)
            else:
                y = decrease_or_increase_numbers(li, operation_number, rounded=rounded, increase=True)
            data.append(y)

    elif float!=False:
        for li in list_of_lists:
            y = list_float(li)
            data.append(y)

    elif integer!=False:
        for li in list_of_lists:
            y = list_int(li)
            data.append(y)

    elif string!=False:
        for li in list_of_lists:
            y = list_str(li)
            data.append(y)

    elif list_rounded!=False:
        for li in list_of_lists:
            y = list_rounded(li)
            data.append(y)

    if list_operator_fn!=False:
        item=operator
        r_num=rounded
        for li in list_of_lists:
            del data[0]
            y = list_operator(li, operation_number, operator=item, rounded=r_num)
            data.append(y)

    return data

def tips():
    """Shortcut tips"""

    print("lists_to_dict = dict(zip(a, [b, c]))")
    print("dict_to_list = dict.values()")

if __name__ == "__main__":
    info()